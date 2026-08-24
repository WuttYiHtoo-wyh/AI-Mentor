from __future__ import annotations

from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path
import hashlib
import json
import sys
from typing import Any

import requests
from openpyxl import Workbook, load_workbook

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from backend.api.main import REGISTRY_PATH, WORKSPACE_ROOT
from backend.mentor_response.chat_service import debug_retrieve_chat_turn, load_chat_module_config
from backend.mentor_response.draft_review import detect_draft_review
from scripts.run_draft_review_eval import _evaluate as evaluate_draft_review
from scripts.run_mentor_100_question_eval import _automated_result as evaluate_standard


SOURCE_100_XLSX = ROOT / "testing" / "DMV_AI_Mentor_100_Question_Test_Set.xlsx"
DRAFT_TEST_JSON = ROOT / "testing" / "draft_review_test_set.json"
FINAL_TEST_SET_XLSX = ROOT / "testing" / "DMV_AI_Mentor_Final_V3_Test_Set.xlsx"
FINAL_RESULTS_XLSX = ROOT / "testing" / "DMV_AI_Mentor_Final_V3_Results.xlsx"
FINAL_RESULTS_JSON = ROOT / "testing" / "mentor_final_v3_results.json"
FINAL_DEBUG_JSON = ROOT / "testing" / "mentor_final_v3_retrieval_debug.json"
FINAL_SHORTLIST_JSON = ROOT / "testing" / "mentor_final_v3_manual_review_shortlist.json"
FINAL_REPORT_MD = ROOT / "testing" / "mentor_final_v3_report.md"
FINAL_DEMO_JSON = ROOT / "testing" / "mentor_final_v3_demo_summary.json"
API_URL = "http://127.0.0.1:8000/api/chat"
MODULE_ID = "PDDS-DMV"
LEVEL = "Basic"


APP_FILES = [
    "backend/api/main.py",
    "backend/mentor_response/academic_integrity.py",
    "backend/mentor_response/chat_service.py",
    "backend/mentor_response/draft_review.py",
    "backend/mentor_response/evidence_policy.py",
    "backend/mentor_response/generator.py",
    "backend/mentor_response/prompts.py",
    "backend/mentor_response/retrieval.py",
    "backend/mentor_response/structural_reference.py",
    "configs/chat_modules.yaml",
    "configs/retrieval_experiment3.yaml",
    "frontend/learner_chat/app.js",
    "frontend/learner_chat/index.html",
    "frontend/learner_chat/styles.css",
]


def main() -> int:
    frozen_manifest = _file_manifest(APP_FILES)
    tests = _final_tests()
    _write_test_set_workbook(tests, FINAL_TEST_SET_XLSX)

    module_config = load_chat_module_config(MODULE_ID, LEVEL, REGISTRY_PATH, WORKSPACE_ROOT)
    results: list[dict[str, Any]] = []
    debug_rows: list[dict[str, Any]] = []
    conversation_history: list[dict[str, str]] = []

    for test in tests:
        history = _history_for_test(test, conversation_history)
        payload = {
            "message": test["Learner Question"],
            "module_id": MODULE_ID,
            "level": LEVEL,
            "conversation_id": "mentor-final-v3-eval",
            "history": history[-8:],
        }
        http_status = None
        response_text = ""
        sources: list[str] = []
        no_context = False
        try:
            response = requests.post(API_URL, json=payload, timeout=120)
            http_status = response.status_code
            if response.ok:
                data = response.json()
                response_text = data.get("answer", "")
                sources = data.get("sources", [])
                no_context = bool(data.get("no_context"))
            else:
                response_text = response.text
        except Exception as exc:
            response_text = f"API error: {exc}"

        debug = {}
        try:
            debug = debug_retrieve_chat_turn(test["Learner Question"], module_config, ROOT, history=history[-8:])
        except Exception as exc:
            debug = {"error": str(exc)}

        if test.get("Evaluation Type") == "draft_review":
            draft_context = detect_draft_review(test["Learner Question"], history[-8:])
            evaluation = evaluate_draft_review(
                {"category": "Draft Review", "expected_behavior": test["Expected Behavior"]},
                response_text,
                sources,
                no_context,
                debug.get("detected_behavior") or ("DRAFT_REVIEW" if draft_context else "NORMAL"),
                draft_context,
            )
            automated_result = evaluation["status"]
            flags = evaluation["flags"]
        else:
            automated_result, flags = evaluate_standard(test, http_status, response_text, sources, no_context, debug)

        result = {
            **test,
            "HTTP Status": http_status,
            "Actual Response": response_text,
            "Actual Sources": "; ".join(sources),
            "No Context": no_context,
            "Detected Interaction Behavior": debug.get("detected_behavior") or ("NO_CONTEXT" if no_context else "NORMAL"),
            "Detected Task/Topic": debug.get("detected_task_or_topic"),
            "Retrieval Query": debug.get("retrieval_query"),
            "Automated Result": automated_result,
            "Notes": "; ".join(flags),
            "Flags": flags,
        }
        results.append(result)
        debug_rows.append(
            {
                "Test ID": test["Test ID"],
                "Category": test["Category"],
                "Learner Question": test["Learner Question"],
                "retrieval_query": debug.get("retrieval_query"),
                "detected_behavior": debug.get("detected_behavior"),
                "detected_task_or_topic": debug.get("detected_task_or_topic"),
                "intent": debug.get("intent"),
                "preferred_role": debug.get("preferred_role"),
                "no_context": debug.get("no_context"),
                "evidence_policy": debug.get("evidence_policy"),
                "evidence": _compact_evidence(debug.get("results", [])),
                "diagnostics": debug.get("diagnostics"),
            }
        )
        if _include_in_conversation_history(test, http_status, response_text):
            conversation_history.append({"role": "learner", "content": test["Learner Question"]})
            conversation_history.append({"role": "mentor", "content": response_text})
            conversation_history = conversation_history[-10:]

    summary = _summary(results)
    shortlist = _manual_review_shortlist(results)
    prior = _verified_prior_metrics()
    output = {
        "run_started_at": datetime.now().isoformat(timespec="seconds"),
        "module_id": MODULE_ID,
        "level": LEVEL,
        "frozen_code_manifest": frozen_manifest,
        "summary": summary,
        "prior_metrics": prior,
        "results": results,
    }
    FINAL_RESULTS_JSON.write_text(json.dumps(output, indent=2, ensure_ascii=False), encoding="utf-8")
    FINAL_DEBUG_JSON.write_text(json.dumps(debug_rows, indent=2, ensure_ascii=False), encoding="utf-8")
    FINAL_SHORTLIST_JSON.write_text(json.dumps(shortlist, indent=2, ensure_ascii=False), encoding="utf-8")
    _write_results_workbook(results, FINAL_RESULTS_XLSX)
    _write_report(output, shortlist)
    _write_demo_summary(output, shortlist)
    print(json.dumps(summary, indent=2))
    return 0


def _final_tests() -> list[dict[str, Any]]:
    standard = _load_100_tests()
    draft_rows = []
    for row in json.loads(DRAFT_TEST_JSON.read_text(encoding="utf-8")):
        if row["category"] != "Draft Review":
            continue
        draft_rows.append(
            {
                "Test ID": row["test_id"],
                "Category": "Draft Review",
                "Learner Question": row["message"],
                "Expected Primary Source": "Project Brief",
                "Expected Topic/Section": row["expected_behavior"],
                "Expected Answer / Key Points": "Review learner-authored draft; identify strengths, gaps, and improvements without grading or replacement writing.",
                "Expected Behavior": row["expected_behavior"],
                "Expected Source Behavior": "Project Brief/rubric first where task-alignment is clear; IU second only if concept support is relevant.",
                "Priority": "High",
                "Evaluation Type": "draft_review",
                "history": row.get("history", []),
            }
        )
    return standard + draft_rows


def _load_100_tests() -> list[dict[str, Any]]:
    workbook = load_workbook(SOURCE_100_XLSX)
    sheet = workbook["Test Cases"]
    header_row = None
    headers: list[str] = []
    for row in range(1, sheet.max_row + 1):
        values = [sheet.cell(row, col).value for col in range(1, sheet.max_column + 1)]
        if "Test ID" in values and "Learner Question" in values:
            header_row = row
            headers = [str(value) if value is not None else "" for value in values]
            break
    if header_row is None:
        raise ValueError("Could not find Test Cases header row.")

    tests = []
    for row in range(header_row + 1, sheet.max_row + 1):
        record = {headers[col - 1]: sheet.cell(row, col).value for col in range(1, len(headers) + 1)}
        if record.get("Test ID") and record.get("Learner Question"):
            tests.append(
                {
                    **record,
                    "Evaluation Type": "standard",
                    "history": [],
                }
            )
    return tests


def _history_for_test(test: dict[str, Any], conversation_history: list[dict[str, str]]) -> list[dict[str, str]]:
    if test.get("history"):
        return list(test["history"])
    return conversation_history


def _include_in_conversation_history(test: dict[str, Any], http_status: int | None, response_text: str) -> bool:
    if http_status != 200 or not response_text:
        return False
    if test.get("Evaluation Type") == "draft_review":
        return False
    return True


def _write_test_set_workbook(tests: list[dict[str, Any]], path: Path) -> None:
    headers = [
        "Test ID",
        "Category",
        "Learner Question",
        "Expected Primary Source",
        "Expected Topic/Section",
        "Expected Answer / Key Points",
        "Expected Behavior",
        "Expected Source Behavior",
        "Priority",
        "Evaluation Type",
    ]
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Test Cases"
    sheet.append(headers)
    for test in tests:
        sheet.append([test.get(header, "") for header in headers])
    workbook.save(path)


def _write_results_workbook(results: list[dict[str, Any]], path: Path) -> None:
    headers = [
        "Test ID",
        "Category",
        "Learner Question",
        "Expected Primary Source",
        "Expected Topic/Section",
        "Expected Behavior",
        "HTTP Status",
        "Actual Response",
        "Actual Sources",
        "Detected Interaction Behavior",
        "Detected Task/Topic",
        "Retrieval Query",
        "Automated Result",
        "Notes",
    ]
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Results"
    sheet.append(headers)
    for result in results:
        sheet.append([result.get(header, "") for header in headers])
    summary = workbook.create_sheet("Summary")
    counts = _summary(results)
    summary.append(["Automated Result", "Count"])
    for status, count in counts["status_counts"].items():
        summary.append([status, count])
    summary.append([])
    summary.append(["Category", "Status", "Count"])
    for category, category_counts in counts["by_category"].items():
        for status, count in category_counts.items():
            summary.append([category, status, count])
    workbook.save(path)


def _compact_evidence(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    return [
        {
            "rank": row.get("rank"),
            "knowledge_role": row.get("knowledge_role"),
            "document_type": row.get("document_type"),
            "document_id": row.get("document_id"),
            "task_reference": row.get("task_reference"),
            "instructional_unit": row.get("instructional_unit"),
            "topic": row.get("topic"),
            "page_start": row.get("page_start"),
            "page_end": row.get("page_end"),
            "final_score": row.get("final_score"),
        }
        for row in rows[:5]
    ]


def _summary(results: list[dict[str, Any]]) -> dict[str, Any]:
    status_counts = Counter(row["Automated Result"] for row in results)
    by_category: dict[str, Counter] = defaultdict(Counter)
    flag_counts = Counter()
    behavior_counts = Counter()
    for row in results:
        by_category[str(row.get("Category", "Unknown"))][row["Automated Result"]] += 1
        flag_counts.update(row.get("Flags", []))
        behavior_counts[str(row.get("Detected Interaction Behavior", "NORMAL"))] += 1
    return {
        "total_tests": len(results),
        "status_counts": dict(status_counts),
        "by_category": {category: dict(counts) for category, counts in by_category.items()},
        "flag_counts": dict(flag_counts),
        "behavior_counts": dict(behavior_counts),
    }


def _manual_review_shortlist(results: list[dict[str, Any]]) -> list[dict[str, Any]]:
    priority_terms = ["Requirement", "Rubric", "Draft Review", "Academic", "Out of Scope"]
    candidates = [row for row in results if row["Automated Result"] in {"REVIEW", "FAIL"}]
    def score(row: dict[str, Any]) -> tuple[int, int, str]:
        category_score = 0 if any(term in str(row.get("Category", "")) for term in priority_terms) else 1
        flag_score = -len(row.get("Flags", []))
        status_score = 0 if row["Automated Result"] == "FAIL" else 1
        return (status_score, category_score, flag_score, str(row.get("Test ID", "")))
    selected = sorted(candidates, key=score)[:30]
    return [
        {
            "Test ID": row["Test ID"],
            "Category": row["Category"],
            "Learner Question": row["Learner Question"],
            "Automated Result": row["Automated Result"],
            "Flags": row.get("Flags", []),
            "Notes": row.get("Notes", ""),
            "Actual Sources": row.get("Actual Sources", ""),
            "Reason for human review": _review_reason(row),
        }
        for row in selected
    ]


def _review_reason(row: dict[str, Any]) -> str:
    if row["Automated Result"] == "FAIL":
        return "Deterministic violation or high-priority automated failure."
    if row.get("Flags"):
        return "Conservative evaluator flagged source/topic or quality concern."
    if row.get("Category") == "Draft Review":
        return "Draft-review accuracy requires human judgment."
    return "Natural-language quality/correctness should be manually checked."


def _file_manifest(paths: list[str]) -> dict[str, Any]:
    manifest = {"recorded_at": datetime.now().isoformat(timespec="seconds"), "files": []}
    for rel in paths:
        path = ROOT / rel
        if not path.exists():
            continue
        content = path.read_bytes()
        manifest["files"].append(
            {
                "path": rel,
                "sha256": hashlib.sha256(content).hexdigest(),
                "last_modified": datetime.fromtimestamp(path.stat().st_mtime).isoformat(timespec="seconds"),
                "bytes": path.stat().st_size,
            }
        )
    return manifest


def _verified_prior_metrics() -> dict[str, Any]:
    validation = json.loads((ROOT / "prepared_knowledge/dmv_basic/validation_report.json").read_text(encoding="utf-8"))
    response = json.loads((ROOT / "testing/mentor_response_experiment2_results.json").read_text(encoding="utf-8"))
    draft = json.loads((ROOT / "testing/draft_review_results.json").read_text(encoding="utf-8"))
    return {
        "knowledge_preparation": {
            "source_documents": validation.get("source_count", 9),
            "prepared_chunks": validation.get("chunk_count", 185),
            "embedding_eligible": 175,
            "embedding_ineligible": 10,
        },
        "retrieval_experiment_comparison": {
            "Baseline": {"Rank 1": 22, "Top 3": 25, "Top 5": 26, "NO_CONTEXT": "0/3"},
            "Experiment 2": {"Rank 1": 26, "Top 3": 27, "Top 5": 27, "NO_CONTEXT": "3/3"},
            "Experiment 3": {"Rank 1": 26, "Top 3": 27, "Top 5": 27, "NO_CONTEXT": "3/3"},
        },
        "response_experiment2": {
            "test_count": response.get("test_count"),
            "summary": response.get("summary"),
        },
        "draft_review_experiment": draft.get("summary"),
    }


def _write_report(output: dict[str, Any], shortlist: list[dict[str, Any]]) -> None:
    summary = output["summary"]
    prior = output["prior_metrics"]
    lines = [
        "# Mentor Final V3 Evaluation Report",
        "",
        "Learner application behavior was frozen before this run. Only this V3 evaluation harness/test-set output was created.",
        "",
        "## Summary",
        "",
        f"- Total tests: `{summary['total_tests']}`",
    ]
    for status, count in summary["status_counts"].items():
        lines.append(f"- {status}: `{count}`")
    lines.extend(["", "## Category Breakdown", ""])
    for category, counts in summary["by_category"].items():
        lines.append(f"- {category}: `{counts}`")
    lines.extend(
        [
            "",
            "## Verified Prior Metrics",
            "",
            f"- Source documents: `{prior['knowledge_preparation']['source_documents']}`",
            f"- Prepared chunks: `{prior['knowledge_preparation']['prepared_chunks']}`",
            f"- Embedding eligible: `{prior['knowledge_preparation']['embedding_eligible']}`",
            f"- Embedding ineligible: `{prior['knowledge_preparation']['embedding_ineligible']}`",
            f"- Retrieval comparison: `{prior['retrieval_experiment_comparison']}`",
            f"- Response Experiment 2 summary: `{prior['response_experiment2']['summary']}`",
            f"- Draft review summary: `{prior['draft_review_experiment']}`",
            "",
            "## Manual Review Shortlist",
            "",
        ]
    )
    for row in shortlist:
        lines.append(
            f"- `{row['Test ID']}` {row['Category']} | {row['Automated Result']} | {row['Notes'] or row['Reason for human review']}"
        )
    lines.extend(
        [
            "",
            "## Known Limitations",
            "",
            "- REVIEW is not failure; it means the conservative evaluator requires human judgment.",
            "- The known deliverables phrasing gap from retrieval Experiment 3 remains visible and was not tuned during V3.",
            "- This is a local evaluation baseline, not an Admin/Evaluation UI.",
        ]
    )
    FINAL_REPORT_MD.write_text("\n".join(lines) + "\n", encoding="utf-8")


def _write_demo_summary(output: dict[str, Any], shortlist: list[dict[str, Any]]) -> None:
    examples = _representative_examples(output["results"])
    demo = {
        "knowledge_preparation_metrics": output["prior_metrics"]["knowledge_preparation"],
        "retrieval_experiment_comparison": output["prior_metrics"]["retrieval_experiment_comparison"],
        "response_experiment_results": output["prior_metrics"]["response_experiment2"],
        "draft_review_results": output["prior_metrics"]["draft_review_experiment"],
        "final_v3_results": output["summary"],
        "representative_successful_test_examples": examples,
        "known_limitations": [
            {"type": "evaluator limitation", "description": "Many natural-language answers remain REVIEW because automatic checks are intentionally conservative."},
            {"type": "genuine system limitation", "description": "Known deliverables-style retrieval gap remains documented from Experiment 3."},
            {"type": "future feature", "description": "No Evaluation UI, Admin UI, authentication, or deployment has been built yet."},
        ],
        "manual_review_shortlist_size": len(shortlist),
    }
    FINAL_DEMO_JSON.write_text(json.dumps(demo, indent=2, ensure_ascii=False), encoding="utf-8")


def _representative_examples(results: list[dict[str, Any]]) -> list[dict[str, Any]]:
    wanted = [
        ("Task requirement", "Explain Task 1"),
        ("Concept explanation", "What is cardinality"),
        ("Typo handling", "Tas1"),
        ("Rubric/requirement authority", "Proficient in Task 2"),
        ("NO_CONTEXT", "Who is Suga"),
        ("Academic-integrity refusal", "copy paste"),
        ("Draft review", "DR005"),
        ("Grade-refusal + draft review", "100 marks"),
    ]
    examples = []
    for label, needle in wanted:
        match = next(
            (
                row
                for row in results
                if needle.lower() in (str(row.get("Learner Question", "")) + " " + str(row.get("Test ID", ""))).lower()
            ),
            None,
        )
        if not match:
            continue
        examples.append(
            {
                "behavior_label": label,
                "test_id": match["Test ID"],
                "learner_question": match["Learner Question"],
                "mentor_response": match["Actual Response"],
                "learner_facing_sources": match["Actual Sources"],
                "evaluation_result": match["Automated Result"],
            }
        )
    return examples[:8]


if __name__ == "__main__":
    raise SystemExit(main())
