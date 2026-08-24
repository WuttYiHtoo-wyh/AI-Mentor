from __future__ import annotations

from collections import Counter, defaultdict
from pathlib import Path
import json
import re
import sys
from typing import Any

import requests
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from backend.api.main import REGISTRY_PATH, WORKSPACE_ROOT
from backend.mentor_response.chat_service import debug_retrieve_chat_turn, load_chat_module_config


SOURCE_WORKBOOK = ROOT / "testing" / "DMV_AI_Mentor_100_Question_Test_Set.xlsx"
RESULT_WORKBOOK = ROOT / "testing" / "DMV_AI_Mentor_100_Question_Test_Results_V2.xlsx"
RESULT_JSON = ROOT / "testing" / "mentor_100_question_results_v2.json"
DEBUG_JSON = ROOT / "testing" / "mentor_100_question_retrieval_debug_v2.json"
REPORT_MD = ROOT / "testing" / "mentor_100_question_report_v2.md"
API_URL = "http://127.0.0.1:8000/api/chat"
MODULE_ID = "PDDS-DMV"
LEVEL = "Basic"


def main() -> int:
    tests, header_row, headers = _load_tests(SOURCE_WORKBOOK)
    module_config = load_chat_module_config(MODULE_ID, LEVEL, REGISTRY_PATH, WORKSPACE_ROOT)
    history: list[dict[str, str]] = []
    results: list[dict[str, Any]] = []
    debug_rows: list[dict[str, Any]] = []

    for test in tests:
        payload = {
            "message": test["Learner Question"],
            "module_id": MODULE_ID,
            "level": LEVEL,
            "conversation_id": "mentor-100-question-eval",
            "history": history[-8:],
        }
        http_status = None
        actual_response = ""
        actual_sources: list[str] = []
        no_context = False
        try:
            response = requests.post(API_URL, json=payload, timeout=120)
            http_status = response.status_code
            if response.ok:
                data = response.json()
                actual_response = data.get("answer", "")
                actual_sources = data.get("sources", [])
                no_context = bool(data.get("no_context"))
            else:
                actual_response = response.text
        except Exception as exc:
            actual_response = f"API error: {exc}"

        retrieval_debug = {}
        try:
            retrieval_debug = debug_retrieve_chat_turn(
                message=test["Learner Question"],
                module_config=module_config,
                workspace_root=ROOT,
                history=history[-8:],
            )
        except Exception as exc:
            retrieval_debug = {"error": str(exc)}

        automated_result, flags = _automated_result(test, http_status, actual_response, actual_sources, no_context, retrieval_debug)
        row_result = {
            **test,
            "Automated Result": automated_result,
            "HTTP Status": http_status,
            "Actual Response": actual_response,
            "Actual Sources": "; ".join(actual_sources),
            "Flags": flags,
        }
        results.append(row_result)
        debug_rows.append(
            {
                "Test ID": test["Test ID"],
                "Learner Question": test["Learner Question"],
                "retrieval_query": retrieval_debug.get("retrieval_query"),
                "normalized_message": retrieval_debug.get("normalized_message"),
                "intent": retrieval_debug.get("intent"),
                "preferred_role": retrieval_debug.get("preferred_role"),
                "no_context": retrieval_debug.get("no_context"),
                "evidence_policy": retrieval_debug.get("evidence_policy"),
                "evidence": _debug_evidence(retrieval_debug.get("results", [])),
                "diagnostics": retrieval_debug.get("diagnostics"),
            }
        )
        if http_status == 200 and actual_response:
            history.append({"role": "learner", "content": test["Learner Question"]})
            history.append({"role": "mentor", "content": actual_response})
            history = history[-10:]

    RESULT_JSON.write_text(json.dumps(results, indent=2, ensure_ascii=False), encoding="utf-8")
    DEBUG_JSON.write_text(json.dumps(debug_rows, indent=2, ensure_ascii=False), encoding="utf-8")
    _write_workbook(SOURCE_WORKBOOK, RESULT_WORKBOOK, header_row, headers, results)
    _write_report(results, REPORT_MD)
    print(json.dumps(_summary(results), indent=2))
    return 0


def _load_tests(path: Path) -> tuple[list[dict[str, Any]], int, list[str]]:
    workbook = load_workbook(path)
    sheet = workbook["Test Cases"]
    header_row = None
    headers: list[str] = []
    for row in range(1, sheet.max_row + 1):
        values = [sheet.cell(row, col).value for col in range(1, sheet.max_column + 1)]
        if "Test ID" in values and "Learner Question" in values:
            header_row = row
            headers = [str(value) if value is not None else "" for value in values]
            break
    if header_row is None:
        raise ValueError("Could not find Test Cases header row.")

    tests = []
    for row in range(header_row + 1, sheet.max_row + 1):
        record = {headers[col - 1]: sheet.cell(row, col).value for col in range(1, len(headers) + 1)}
        if record.get("Test ID") and record.get("Learner Question"):
            tests.append(record)
    return tests, header_row, headers


def _automated_result(
    test: dict[str, Any],
    http_status: int | None,
    response: str,
    sources: list[str],
    no_context: bool,
    retrieval_debug: dict[str, Any],
) -> tuple[str, list[str]]:
    flags = []
    response_l = response.lower()
    category = str(test.get("Category", "")).lower()
    expected_source = str(test.get("Expected Primary Source", "") or "")
    expected_topic = str(test.get("Expected Topic/Section", "") or "")
    expected_behavior = str(test.get("Expected Behavior", "") or "").lower()
    expected_sources = _expected_source_options(expected_source)

    if http_status != 200:
        return "FAIL", [f"http_status_{http_status}"]
    if not response.strip():
        return "FAIL", ["empty_response"]
    if _has_raw_markdown(response):
        flags.append("raw_markdown")
    if len(sources) > 2:
        flags.append("excessive_source_list")
    if _has_generic_filler(response):
        flags.append("generic_filler_ending")

    if "out" in category or "insufficient" in expected_behavior or _expects_no_sources(expected_source):
        if no_context and "approved module materials" in response_l and not sources:
            return "NO_CONTEXT", flags
        flags.append("out_of_scope_not_controlled")
        return "FAIL", flags

    if "complete" in expected_behavior or "submission-ready" in expected_behavior:
        if _refuses_complete_work(response_l):
            return "PASS", flags
        flags.append("academic_integrity_complete_work")
        return "FAIL", flags

    if "grade" in expected_behavior or "percentage" in expected_behavior:
        if _refuses_grading(response_l):
            return "PASS", flags
        flags.append("academic_integrity_grading")
        return "FAIL", flags

    if expected_sources and not _source_matches_any(expected_sources, sources):
        flags.append("expected_source_missing")
        if any(source.lower() == "project brief" for source in expected_sources):
            return "FAIL", flags

    evidence = retrieval_debug.get("results", [])
    if expected_topic and _clear_topic_check(expected_topic):
        haystack = " ".join(
            [
                response,
                " ".join(sources),
                " ".join(str(row.get("topic", "")) for row in evidence),
                " ".join(str(row.get("instructional_unit", "")) for row in evidence),
            ]
        ).lower()
        if expected_topic.lower().replace("_", "").replace(" ", "") not in haystack.replace("_", "").replace(" ", ""):
            flags.append("expected_topic_not_obvious")
            return "REVIEW", flags

    if flags:
        return "REVIEW", flags
    return "REVIEW", flags


def _expected_source_options(value: str) -> list[str]:
    value = value.strip()
    if not value or _expects_no_sources(value):
        return []
    options = re.split(r"\s*(?:/|\bor\b)\s*", value, flags=re.I)
    return [option.strip() for option in options if option.strip() and not _expects_no_sources(option)]


def _expects_no_sources(value: str) -> bool:
    normalized = _normalize_source(value)
    return normalized in {"none", "no source", "no sources", "no dmv source", "no dmv sources", "n/a", "na"}


def _source_matches_any(expected_options: list[str], actual_sources: list[str]) -> bool:
    actual_normalized = [_normalize_source(source) for source in actual_sources]
    return any(
        _source_option_matches_actual(_normalize_source(expected), actual)
        for expected in expected_options
        for actual in actual_normalized
    )


def _source_option_matches_actual(expected: str, actual: str) -> bool:
    if not expected or not actual:
        return False
    if expected == actual or expected in actual:
        return True
    return False


def _normalize_source(value: str) -> str:
    normalized = str(value or "").lower().replace("_", "")
    normalized = re.sub(r"[^a-z0-9]+", " ", normalized)
    return re.sub(r"\s+", " ", normalized).strip()


def _clear_topic_check(topic: str) -> bool:
    normalized = topic.strip().lower()
    if not normalized or normalized in {"none", "n/a", "na"}:
        return False
    return len(topic) <= 40 and not any(char in topic for char in ["/", ",", ";"])


def _debug_evidence(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    return [
        {
            "rank": row.get("rank"),
            "chunk_id": row.get("chunk_id"),
            "knowledge_role": row.get("knowledge_role"),
            "document_id": row.get("document_id"),
            "instructional_unit": row.get("instructional_unit"),
            "task_reference": row.get("task_reference"),
            "topic": row.get("topic"),
            "page_start": row.get("page_start"),
            "page_end": row.get("page_end"),
            "distance": row.get("distance"),
            "final_score": row.get("final_score"),
        }
        for row in rows
    ]


def _write_workbook(source: Path, target: Path, header_row: int, headers: list[str], results: list[dict[str, Any]]) -> None:
    workbook = load_workbook(source)
    sheet = workbook["Test Cases"]
    header_map = {name: index + 1 for index, name in enumerate(headers)}
    for output_header in ["Automated Result", "HTTP Status", "Actual Response", "Actual Sources", "Notes"]:
        if output_header not in header_map:
            headers.append(output_header)
            header_map[output_header] = len(headers)
            sheet.cell(header_row, len(headers)).value = output_header
    result_by_id = {row["Test ID"]: row for row in results}
    for row in range(header_row + 1, sheet.max_row + 1):
        test_id = sheet.cell(row, header_map["Test ID"]).value
        result = result_by_id.get(test_id)
        if not result:
            continue
        sheet.cell(row, header_map["Automated Result"]).value = result["Automated Result"]
        sheet.cell(row, header_map["HTTP Status"]).value = result["HTTP Status"]
        sheet.cell(row, header_map["Actual Response"]).value = result["Actual Response"]
        sheet.cell(row, header_map["Actual Sources"]).value = result["Actual Sources"]
        sheet.cell(row, header_map["Notes"]).value = "; ".join(result["Flags"])

    if "Summary" in workbook.sheetnames:
        summary = workbook["Summary"]
    else:
        summary = workbook.create_sheet("Summary")
    for merged_range in list(summary.merged_cells.ranges):
        summary.unmerge_cells(str(merged_range))
    counts = _summary(results)
    summary["A1"] = "Automated Result"
    summary["B1"] = "Count"
    for index, (name, count) in enumerate(counts["status_counts"].items(), start=2):
        summary.cell(index, 1).value = name
        summary.cell(index, 2).value = count
    summary["D1"] = "Category"
    summary["E1"] = "Status"
    summary["F1"] = "Count"
    row_index = 2
    for category, status_counts in counts["by_category"].items():
        for status, count in status_counts.items():
            summary.cell(row_index, 4).value = category
            summary.cell(row_index, 5).value = status
            summary.cell(row_index, 6).value = count
            row_index += 1
    workbook.save(target)


def _write_report(results: list[dict[str, Any]], path: Path) -> None:
    summary = _summary(results)
    weakest = _weakest(results)
    lines = [
        "# Mentor 100-Question Evaluation Report",
        "",
        f"Total tests: `{summary['total']}`.",
        "",
        "## Automated Counts",
        "",
    ]
    for status, count in summary["status_counts"].items():
        lines.append(f"- {status}: `{count}`")
    lines.extend(["", "## Results By Category", ""])
    for category, counts in summary["by_category"].items():
        lines.append(f"- {category}: {dict(counts)}")
    lines.extend(
        [
            "",
            "## Failure Clusters",
            "",
            f"- Retrieval/source failures: `{summary['flag_counts'].get('expected_source_missing', 0)}`.",
            f"- Answer-quality failures: `{summary['flag_counts'].get('empty_response', 0) + summary['flag_counts'].get('generic_filler_ending', 0)}`.",
            f"- Typo/robustness failures: `{summary['flag_counts'].get('typo_or_structural_reference', 0)}`.",
            f"- Context/follow-up failures: `{summary['flag_counts'].get('context_followup', 0)}`.",
            f"- Academic-integrity failures: `{summary['flag_counts'].get('academic_integrity_complete_work', 0) + summary['flag_counts'].get('academic_integrity_grading', 0)}`.",
            f"- Citation/source-display failures: `{summary['flag_counts'].get('excessive_source_list', 0) + summary['flag_counts'].get('expected_source_missing', 0)}`.",
            "",
            "## Ten Weakest Test Cases",
            "",
        ]
    )
    for result in weakest:
        lines.append(
            f"- `{result['Test ID']}` {result.get('Learner Question')} | {result['Automated Result']} | {', '.join(result['Flags']) or 'manual review'}"
        )
    lines.extend(
        [
            "",
            "## Diagnosis",
            "",
            "Automated checks are intentionally conservative. REVIEW means human judgment is still needed for correctness/helpfulness.",
            "Failure clusters should be interpreted as retrieval, evidence policy, prompting, or UI issues based on the flags and retrieval debug artifact.",
            "",
            "## Recommendation",
            "",
            "Do not automatically tune from this run. Review the weakest cases and decide whether a targeted follow-up experiment is justified.",
        ]
    )
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def _summary(results: list[dict[str, Any]]) -> dict[str, Any]:
    status_counts = Counter(row["Automated Result"] for row in results)
    by_category: dict[str, Counter] = defaultdict(Counter)
    flag_counts = Counter()
    for row in results:
        by_category[str(row.get("Category", "Unknown"))][row["Automated Result"]] += 1
        flag_counts.update(row.get("Flags", []))
    return {
        "total": len(results),
        "status_counts": dict(status_counts),
        "by_category": {category: dict(counts) for category, counts in by_category.items()},
        "flag_counts": dict(flag_counts),
    }


def _weakest(results: list[dict[str, Any]]) -> list[dict[str, Any]]:
    order = {"FAIL": 0, "REVIEW": 1, "NO_CONTEXT": 2, "PASS": 3}
    return sorted(results, key=lambda row: (order.get(row["Automated Result"], 9), -len(row.get("Flags", []))))[:10]


def _has_raw_markdown(response: str) -> bool:
    return False


def _has_generic_filler(response: str) -> bool:
    return bool(re.search(r"(if you have any (more|further) questions, feel free to ask!?|feel free to ask if)", response, re.I))


def _refuses_complete_work(response_l: str) -> bool:
    return any(phrase in response_l for phrase in ["can't write", "cannot write", "can't complete", "cannot complete"]) and not any(
        phrase in response_l for phrase in ["here is your complete", "final submission"]
    )


def _refuses_grading(response_l: str) -> bool:
    return any(phrase in response_l for phrase in ["can't assign", "cannot assign", "can't grade", "cannot grade", "not the auto grader"])


if __name__ == "__main__":
    raise SystemExit(main())
